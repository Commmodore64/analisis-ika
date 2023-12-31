import asyncio
import hashlib
import re
from array import array
from datetime import datetime, timedelta
from functools import wraps

import jwt
import mysql.connector
import numpy as np
from flask import (Flask, jsonify, make_response, render_template, request,
                   session)
from flask_sqlalchemy import SQLAlchemy
from numpy import ndarray
from openpyxl import load_workbook

#importacion de algoritmos
from pso import ejecutar_pso
from topsispso import ejecutar_topsispso

db = SQLAlchemy()
app = Flask(__name__)
app.config["SQLALCHEMY_DATABASE_URI"] = "postgresql://postgres:53Zc56erWZsY7@db.ogyulszumjcstdztensn.supabase.co:5432/postgres"
app.config['SECRET_KEY'] = '563cebb3aceb49e0a6c79ded5c717235'

db.init_app(app)
class SupaUser(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombredb = db.Column(db.String, unique=True, nullable=False)
    escolaridaddb = db.Column(db.String, nullable=False)
    nacimientodb = db.Column(db.Date, nullable=False)
    emaildb = db.Column(db.String, nullable=False)
    passworddb = db.Column(db.String, nullable=False)
    
    

@app.route('/pso')
def pso():
    try:
        # Obtén los datos del formulario
        w_input = [request.form.get(f'w[{i}]', '') for i in range(5)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        wwi = float(request.form['wwi'])
        c1 = float(request.form['c1'])
        c2 = float(request.form['c2'])
        T = int(request.form['T'])
        r1_input = request.form['r1']
        r2_input = request.form['r2']
        r1 = [float(num.strip()) for num in r1_input.split(',')]
        r2 = [float(num.strip()) for num in r2_input.split(',')]
        
        # Llama a la función de procesar_datos en pso.py
        nuevosDatos = asyncio.run(ejecutar_pso(w, wwi, c1, c2, T, r1, r2))

        return render_template('pso.html', nuevosDatos=nuevosDatos)
    except Exception as e:
        return render_template('pso.html', error_message=str(e))


@app.route('/pso', methods=['POST'])
def calcular_pso():
    try:
        # Obtén los datos del formulario
        w_input = [request.form.get(f'w[{i}]', '') for i in range(5)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        wwi = float(request.form['wwi'])
        c1 = float(request.form['c1'])
        c2 = float(request.form['c2'])
        T = int(request.form['T'])
        # Divide las cadenas de texto en listas
        r1_input = request.form['r1']
        r2_input = request.form['r2']
        r1 = [float(num.strip()) for num in r1_input.split(',')]
        r2 = [float(num.strip()) for num in r2_input.split(',')]

        # Llama a la función de PSO en pso.py
        nuevosDatos = asyncio.run(ejecutar_pso(w, wwi, c1, c2, T, r1, r2))
        print("Resultados de la ejecución:", nuevosDatos)

        # Obtén los resultados específicos que deseas mostrar
        # dataGBP = resultados['dataGBP']
        # dataGBF = resultados['dataGBF']
        # dataResult = resultados['dataResult']

        # Puedes hacer lo que quieras con los resultados, por ejemplo, pasarlos al template
        return jsonify(nuevosDatos)
    except Exception as e:
        # Manejo de errores, por ejemplo, mostrar un mensaje de error en la interfaz
       print(f'Error en calcular_pso: {str(e)}')
    return jsonify({'error': 'Ocurrió un error en el servidor'}), 500

@app.route('/moorapso', methods=['POST'])
def calcular_moorapso():
    try:
        # Obtén los datos del formulario
        w_input = [request.form.get(f'w[{i}]', '') for i in range(5)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        wwi = float(request.form['wwi'])
        c1 = float(request.form['c1'])
        c2 = float(request.form['c2'])
        T = int(request.form['T'])
        # Divide las cadenas de texto en listas
        r1_input = request.form['r1']
        r2_input = request.form['r2']
        r1 = [float(num.strip()) for num in r1_input.split(',')]
        r2 = [float(num.strip()) for num in r2_input.split(',')]

        # Llama a la función de PSO en pso.py
        nuevosDatos = asyncio.run(ejecutar_pso(w, wwi, c1, c2, T, r1, r2))
        print("Resultados de la ejecución:", nuevosDatos)

        # Obtén los resultados específicos que deseas mostrar
        # dataGBP = resultados['dataGBP']
        # dataGBF = resultados['dataGBF']
        # dataResult = resultados['dataResult']

        # Puedes hacer lo que quieras con los resultados, por ejemplo, pasarlos al template
        return jsonify(nuevosDatos)
    except Exception as e:
        # Manejo de errores, por ejemplo, mostrar un mensaje de error en la interfaz
       print(f'Error en calcular_pso: {str(e)}')
    return jsonify({'error': 'Ocurrió un error en el servidor'}), 500


@app.route('/moorapso')
def moorapso():
    return render_template('moorapso.html')

@app.route('/topsispso')
def topsispso():
    return render_template('topsispso.html')

@app.route('/topsispso', methods=['POST'])
def calcular_topsispso():
    # Obtén los datos del formulario
    wwi = float(request.form['w'])
    c1 = float(request.form['c1'])
    c2 = float(request.form['c2'])
    T = int(request.form['T'])
    # Divide las cadenas de texto en listas
    r1 = [float(val.strip()) for val in request.form['r1'].split(',')]
    r2 = [float(val.strip()) for val in request.form['r2'].split(',')]

    # Llama a la función de PSO en pso.py
    resultados = ejecutar_topsispso(wwi, c1, c2, T, r1, r2)
    # Obtén los resultados específicos que deseas mostrar
    dataGBP = resultados['dataGBP']
    dataGBF = resultados['dataGBF']
    dataResult = resultados['dataResult']
    # Puedes hacer lo que quieras con los resultados, por ejemplo, pasarlos al template
    return render_template('topsispso.html', dataGBP=dataGBP, dataGBF=dataGBF, dataResult=dataResult)


@app.route('/index', methods=['POST'])
def index():
    # Parámetros de control (ingresan)
    w=request.form['w']
    #w=0.3
    # n=int(request.form['n'])  #partículas
    n=9
    c1=request.form['c1']
    #c1=0.50
    c2=request.form['c2']
    #c2=0.50
    e=int(request.form['e'])
    itera=e  #iteraciones

#Función objetivo
# Ri= (t^-i)/ (t^+i+t^-i) , i=1,...,m
# vector negativo / (vector positivo + vector negarivo)
# R2/(R2+R1)


    # Variables aleatorias, estas van a ser ingresadas(posteriormente)
    #R1 --> es el Vector positivo
    FILE_PATH='r1.xlsx'
    SHEET='Hoja1'
    workbook=load_workbook(FILE_PATH,read_only=True)
    sheet=workbook[SHEET]
    r1=[]
    for row in sheet.iter_rows():
        #print(row[0].value)
        r1.append(row[0].value)
        #print(r1)

    #R2 --> es el vector negativo
    #SHEET = 'Hoja1'
    FILE_PATH='r2.xlsx'
    workbook=load_workbook(FILE_PATH,read_only=True)
    sheet=workbook[SHEET]
    r2=[]
    for row in sheet.iter_rows():
        #print(row[0].value)
        r2.append(row[0].value)
        #print(r1)

    # ----------
    # Inicialización de la partícula del enjambre
    # Iniciando cálculo de las iteraciones
    print("")
    print("----")
    print("Iteración # 1")
    CP=[]
    for i in range(n):
        r11=float(r1[i])
        CP.append(float(format(10*(r11-0.5), '.4f')))
        #print("CP=",CP)
    #print("   CP=",CP)
    print("   Posición actual:       CP=",CP)

    # Inicialización de la velocidad
    V=[]
    for i in range(n):
        r21=float(r2[i])
        V.append(float(format((r21-0.5), '.4f')))
    #print("   V=",V)
    #print("   Velocidad actual:             V=",V)

    # Óptimo actual CF(1)
    # Ri= (t^-i)/ (t^+i+t^-i) , i=1,...,m
    # vector negativo / (vector positivo + vector negarivo)
    # R2/(R2+R1)
    CF=[]
    for i in range(n):
        #R2/(R1+R2)
        CForm1=float(r1[i])
        CForm2=float(r2[i])
        CForm3=(float((CForm2)/(CForm2+CForm1)))
        CForm=(format(float(CForm3), '.4f'))
        CF.append(float(CForm))
    print("   Óptimo actual:         CF=",CF)
    #print("   CF=",CF)

    # mejor posición actual
    LBP=[]
    for i in range(n):
        # LBP[1]=CP[1]
        LBPP=format(float(CP[i]), '.4f')
        LBP.append(float(LBPP))
    #print("   LBP=",LBP)
    print("   Mejor posición local:  LBP=",LBP)

    # mejor óptimo local
    LBF=[]
    for i in range(n):
        # LBF[1]=CF[1]
        LBFF=format(float(CF[i]), '.4f')
        LBF.append(float(LBFF))
    #print("   LBF=",LBF)
    print("   Mejor óptimo local:    LBF=",LBF)

    # MEJOR ÓPTIMO GLOBAL
    GBF=[]
    GBF.append(max(LBF))
    GBFt=max(LBF)
    #print("   GBF=", GBF)
    print("   Mejor óptimo global:   GBF=", GBF)

    # MEJOR POSICION GLOBAL
    GBP=[]
    GBF_index=LBF.index(GBFt)
    GBPt=float(LBP[GBF_index])
    GBP.append(LBP[GBF_index])
    #print("   GBP=",GBP)
    print("   Mejor posición global: GBP=",GBP)

    ###########################################
    #  Iteráción 2 a la n
    ##########################################
    it=2
    e=e-2

    while (e>=0):
        print("--------")
        print("Iteración #", it)
        long_V1=len(V)-n
        long_LBP1=len(LBP)-n
        long_CP1=len(CP)-n
        long_GBP1=len(GBP)-1
        for i in range(n):
            # V(i+1)= W*V(i) +c1*r1*(LBP(i)-CP(i))+c2*r2*(GBP)i)-CP(i))
            V1=format(float(w)*float(V[long_V1]), '.4f')
            V2=format(float(c1)*float(r1[i]), '.4f')
            V3=format(float(LBP[long_LBP1])-float(CP[long_CP1]), '.4f')
            V4=format(float(c2)*float(r2[i]), '.4f')
            V5=format(float(GBP[long_GBP1]), '.4f')
            V6=format(float(CP[long_CP1]), '.4f')
            V7=float(V5)-float(V6)
            Vx=format((float(V1)+float(V2)*float(V3)+float(V4)*float(V7)), '.4f')
            V.append(Vx)
            long_V1=long_V1+1
            long_LBP1=long_LBP1+1
            long_CP1=long_CP1+1
        long_V2=len(V)-n
        long_CP2=len(CP)-n

        #CP=CP(i)+V(i)
        for i in range(n):
            CPI=(format(float(CP[long_CP2])+float(V[long_V2]), '.4f'))
            CP.append(float(CPI))
            long_V2=long_V2+1
            long_CP2=long_CP2+1

        #óptimo actual CF
        long_CP1=len(CP)-(2*n)
        #print("long_CP2",long_CP2)
        long_CP2=len(CP)-n
        #print("long_CP3",long_CP3)
        for i in range(n):
            #CF2/(CF2+CF1)
            #print("CP",CP)
            CF1=CP[long_CP1]
            #print("CF1",CF1)
            CF2=CP[long_CP2]
            #print("CF2",CF2)
            CF3=float(CF2+CF1)
            CF12=format(float((CF2)/(CF3)),'.4f')
            CF.append(CF12)
            #print("CF12 ",CF12)
            long_CP1=long_CP1+1
            long_CP2=long_CP2+1

        # mejor óptimo local
        long_CF2=len(CF)-n
        long_LBF1=len(LBF)-(n)
        for i in range(n):
            #Max( CF[i],LBF[i-1])
            #print("CF",CF)
            #print("LBF",LBF)

            CFt=float(CF[long_CF2])
            #print("long_CF2",long_CF2)
            #print("CFt",CFt)
            
            LBFt=float(LBF[long_LBF1])
            #print("long_LBF1",long_LBF1)
            #print("LBFt",LBFt)
            if CFt>LBFt:
                LBF.append(CFt)
            else:
                LBF.append(LBFt)
            #print("LBF",LBF)
            long_CF2=long_CF2+1
            long_LBF1=long_LBF1+1

        # mejor posición actual
        long_CP4=len(CP)-n
        for i in range(n):
            #LBP[i]= posición de LBF[i]-CP[i]
            LBPt=(format(float(CP[long_CP4]), '.4f'))
            LBP.append(float(LBPt))
            long_CP4=long_CP4+1

        # MEJOR ÓPTIMO GLOBAL
        long_LBF=len(CP)-n
        temporal=[]
        for i in range(n):
            temporal.append(LBF[long_LBF])
            long_LBF=long_LBF+1
        GBF.append(max(temporal))
        GBFt=max(temporal)
        #print("GBFt=",GBFt)
        #print("MaxGlobal=",GBFt)

        # MEJOR POSICION GLOBAL
        #print("PosiciónGBP posición en CP")
        #print("CP=",CP)
        #long_GBF=len(CP)-n
        long_GBF=len(CP)-1
        #print("LBF",LBF)
        #print("long_CP=",long_GBF)
        GBPt=CP[long_GBF]
        #print("GBPt",GBPt)
        GBP.append(GBPt)


    # impresión de datos por iteración
        V_imp = len(V)-n
        CP_imp = len(CP)-n
        CF_imp = len(CF)-n
        LBF_imp = len(LBF)-n
        LBP_imp = len(LBP)-n
        GBF_imp = len(GBF)-1
        GBP_imp = len(GBP)-1

        V_impT = []
        for i in range(n):
            V_impT.append(float(V[V_imp]))
            V_imp = V_imp+1
        print("   V=",V_impT)

        CP_impT = []
        for i in range(n):
            CP_impT.append(float(CP[CP_imp]))
            CP_imp = CP_imp+1
        print("   CP=",CP_impT)

        CF_impT = []
        for i in range(n):
            CF_impT.append(float(CF[CF_imp]))
            CF_imp = CF_imp+1
        print("   CF=",CF_impT)

        LBF_impT = []
        for i in range(n):
            LBF_impT.append(float(LBF[LBF_imp]))
            LBF_imp = LBF_imp+1
        print("   LBF=",LBF_impT)

        LBP_impT = []
        for i in range(n):
            LBP_impT.append(float(LBP[LBP_imp]))
            LBP_imp = LBP_imp+1
        print("   LBP=",LBP_impT)

        for i in range(n):
            GBF_impT = (float(GBF[GBF_imp]))
        print("   GBF=",GBF_impT)

        for i in range(n):
            GBP_impT = (float(GBP[GBP_imp]))
        print("   GBP=",GBP_impT)

        e = e-1
        it = it+1
        print("")

    print("*******************")
    GBF_SF = len(GBF)-1
    GBP_SF = len(GBP)-1

    GBF_FIN = (float(GBF[GBF_SF]))
    print("   Mejor posición=",GBF_FIN)

    GBP_FIN = (float(GBP[GBP_SF]))
    print("   Mejor óptimo=",GBP_FIN)
    print("*******************")
    print("")

    context = {
        "c1": c1,
        "c2": c2,
        "w": w,
        "n": n,
        "e": e,
        "itera": itera,
        #"var1": var1,
        #"var2": var2,
        #"var3": var3,

        "r1": r1,
        "r2": r2,

        "V_impT": V_impT,
        "CP_impT": CP_impT,
        "CF_impT": CF_impT,
        "LBF_impT": LBF_impT,
        "LBP_impT": LBP_impT,
        "GBF_impT": GBF_impT,
        "GBP_impT": GBP_impT,

        "GBF_FIN": GBF_FIN,
        "GBP_FIN": GBP_FIN,
    }

    return render_template('index.html', **context)

@app.route('/')
def home():
    # if not session.get('logged_in'):
    #     return render_template('login.html')
    # else:
        return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def log_in():
    if request.method=='POST':
        emailLogin = request.form['email']
        passLogin = request.form['password']
        mensaje = ''
        password_Object = hashlib.sha256(passLogin.encode())
        passwordHashed =  password_Object.hexdigest()
        all_users = SupaUser.query.all()
        for user in all_users:
            print(f"ID: {user.id}, Username: {user.nombredb}, Email: {user.emaildb}, Password: {user.passworddb}")
        if emailLogin == user.emaildb and passwordHashed == user.passworddb:
            print("Usuario Correcto")
            mensaje = "Usuario Correcto"
            session['logged_in'] = True
            return render_template('index.html')
        else:
            print("Usuario Incorrecto")
            mensaje = "Usuario Incorrecto"
            return render_template('login.html')
    else:
        return render_template('login.html')

@app.route('/signup')
def sign_up():
    return render_template('signup.html')

@app.route('/signup', methods=['POST'])
def sign_validation():
    
    listaErrores = {
        'nombre': '',
        'escolaridad' : '',
        'nacimiento' : '',
        'email' : '',
        'password' : ''
    }
    if request.method == 'POST': 
        nombre = request.form['nombre']
        escolaridad = request.form['escolaridad']
        nacimiento = request.form['nacimiento']
        email = request.form['email']
        password = request.form['password']
        mensaje = ''
            
        if not nombre:
            listaErrores['nombre'] = 'Completa este campo'
        
        elif not re.match("^^[a-zA-Z\s]+$", nombre):
                listaErrores['nombre'] = 'Favor de ingresar solo letras'
        
        if not escolaridad:
            listaErrores['escolaridad'] = 'Completa este campo'
        
        elif not re.match("^[a-zA-Z\s]+$", escolaridad):
                listaErrores['escolaridad'] = 'Favor de ingresar solo letras'
        
        if not nacimiento:
            listaErrores['nacimiento'] = 'Completa este campo'

        if not email:
            listaErrores['email'] = 'Completa este campo'
        
        elif not re.match(r'^[\w\.-]+@[\w\.-]+$', email):
                listaErrores['email'] = 'Ingresa un correo valido'
                
        if not password:
            listaErrores['password'] = 'Completa este campo'
                        
        elif len(password) < 8 or len(password) > 12 :
                listaErrores['password'] = 'La contrasena debe tener entre 8 y 12 caracteres' 
        
        if all(value == '' for value in listaErrores.values()):
            password_Object = hashlib.sha256(password.encode())
            passwordHashed =  password_Object.hexdigest()
            try:
                with app.app_context():
                    db.create_all()
                    user = SupaUser(nombredb=nombre, escolaridaddb=escolaridad, nacimientodb=nacimiento, emaildb=email, passworddb=passwordHashed)
                    db.session.add(user)
                    db.session.commit()
                    mensaje = "Registro realizado con éxito!"
                return render_template("/login.html", mensaje=mensaje)
            except Exception as e:
                print(f'Error en la base de datos: {str(e)}')
                mensaje = "Ocurrió un error en la base de datos, por favor inténtalo de nuevo."
        else:
            return render_template("/signup.html", listaErrores=listaErrores)

    return render_template("/signup.html", listaErrores=listaErrores)

if '__main__' == __name__:
    app.run(port=5000, debug=True)
